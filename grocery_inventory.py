"""
Grocery Inventory Management System
Simple program to manage a home grocery inventory list with Excel export functionality.
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class GroceryInventory:
    """Manage a grocery inventory list."""
    
    def __init__(self, data_file="inventory.json"):
        """Initialize the inventory with a JSON data file."""
        self.data_file = data_file
        self.inventory = self.load_inventory()
    
    def load_inventory(self):
        """Load inventory from JSON file."""
        if os.path.exists(self.data_file):
            with open(self.data_file, 'r') as f:
                return json.load(f)
        return []
    
    def save_inventory(self):
        """Save inventory to JSON file."""
        with open(self.data_file, 'w') as f:
            json.dump(self.inventory, f, indent=2)
    
    def add_item(self, name, quantity, unit, category, expiry_date=None):
        """Add a new item to inventory."""
        item = {
            "id": len(self.inventory) + 1,
            "name": name,
            "quantity": quantity,
            "unit": unit,
            "category": category,
            "expiry_date": expiry_date,
            "added_date": datetime.now().strftime("%Y-%m-%d")
        }
        self.inventory.append(item)
        self.save_inventory()
        print(f"✓ Added: {name} ({quantity} {unit})")
    
    def remove_item(self, item_id):
        """Remove an item from inventory by ID."""
        self.inventory = [item for item in self.inventory if item["id"] != item_id]
        self.save_inventory()
        print(f"✓ Removed item with ID: {item_id}")
    
    def update_item(self, item_id, **kwargs):
        """Update an item's details."""
        for item in self.inventory:
            if item["id"] == item_id:
                for key, value in kwargs.items():
                    if key in item:
                        item[key] = value
                self.save_inventory()
                print(f"✓ Updated item ID {item_id}")
                return
        print(f"✗ Item with ID {item_id} not found")
    
    def view_inventory(self):
        """Display all items in inventory."""
        if not self.inventory:
            print("Inventory is empty!")
            return
        
        print("\n" + "="*80)
        print(f"{'ID':<5} {'Item':<20} {'Quantity':<12} {'Category':<15} {'Expiry Date':<15}")
        print("="*80)
        
        for item in self.inventory:
            expiry = item.get("expiry_date", "N/A")
            print(f"{item['id']:<5} {item['name']:<20} {item['quantity']} {item['unit']:<10} {item['category']:<15} {expiry:<15}")
        
        print("="*80 + "\n")
    
    def get_by_category(self, category):
        """Get all items in a specific category."""
        return [item for item in self.inventory if item["category"].lower() == category.lower()]
    
    def export_to_excel(self, filename="grocery_inventory.xlsx"):
        """Export inventory to Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        # Define header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        headers = ["ID", "Item Name", "Quantity", "Unit", "Category", "Expiry Date", "Added Date"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Add data rows
        for item in self.inventory:
            ws.append([
                item["id"],
                item["name"],
                item["quantity"],
                item["unit"],
                item["category"],
                item.get("expiry_date", ""),
                item["added_date"]
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        # Save file
        wb.save(filename)
        print(f"✓ Inventory exported to {filename}")


def main():
    """Main program loop."""
    inventory = GroceryInventory()
    
    while True:
        print("\n--- Grocery Inventory Manager ---")
        print("1. View Inventory")
        print("2. Add Item")
        print("3. Remove Item")
        print("4. Update Item")
        print("5. View by Category")
        print("6. Export to Excel")
        print("7. Exit")
        
        choice = input("\nSelect an option (1-7): ").strip()
        
        if choice == "1":
            inventory.view_inventory()
        
        elif choice == "2":
            name = input("Item name: ").strip()
            quantity = int(input("Quantity: "))
            unit = input("Unit (e.g., kg, liter, pieces): ").strip()
            category = input("Category (e.g., Fruits, Vegetables, Dairy): ").strip()
            expiry = input("Expiry date (YYYY-MM-DD) [optional, press Enter to skip]: ").strip()
            expiry_date = expiry if expiry else None
            inventory.add_item(name, quantity, unit, category, expiry_date)
        
        elif choice == "3":
            inventory.view_inventory()
            try:
                item_id = int(input("Enter item ID to remove: "))
                inventory.remove_item(item_id)
            except ValueError:
                print("✗ Invalid ID")
        
        elif choice == "4":
            inventory.view_inventory()
            try:
                item_id = int(input("Enter item ID to update: "))
                print("What would you like to update?")
                print("1. Quantity  2. Expiry Date")
                update_choice = input("Select (1-2): ").strip()
                
                if update_choice == "1":
                    new_quantity = int(input("New quantity: "))
                    inventory.update_item(item_id, quantity=new_quantity)
                elif update_choice == "2":
                    new_expiry = input("New expiry date (YYYY-MM-DD): ").strip()
                    inventory.update_item(item_id, expiry_date=new_expiry)
            except ValueError:
                print("✗ Invalid input")
        
        elif choice == "5":
            category = input("Enter category name: ").strip()
            items = inventory.get_by_category(category)
            if items:
                print(f"\n--- Items in '{category}' ---")
                for item in items:
                    print(f"  • {item['name']}: {item['quantity']} {item['unit']}")
            else:
                print(f"No items found in category '{category}'")
        
        elif choice == "6":
            filename = input("Enter filename (default: grocery_inventory.xlsx): ").strip()
            if not filename:
                filename = "grocery_inventory.xlsx"
            inventory.export_to_excel(filename)
        
        elif choice == "7":
            print("Thank you for using Grocery Inventory Manager!")
            break
        
        else:
            print("✗ Invalid choice. Please try again.")


if __name__ == "__main__":
    main()
