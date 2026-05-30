"""
Web-based Grocery Inventory Management System
Flask application to manage grocery inventory with Excel export functionality.
"""

from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from datetime import datetime
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io


app = Flask(__name__)
app.config['JSON_SORT_KEYS'] = False
app.secret_key = 'your_secret_key_here'

# Data file path
DATA_FILE = "inventory.json"


class GroceryInventory:
    """Manage a grocery inventory list."""
    
    def __init__(self, data_file=DATA_FILE):
        """Initialize the inventory with a JSON data file."""
        self.data_file = data_file
        self.inventory = self.load_inventory()
    
    def load_inventory(self):
        """Load inventory from JSON file."""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r') as f:
                    return json.load(f)
            except:
                return []
        return []
    
    def save_inventory(self):
        """Save inventory to JSON file."""
        with open(self.data_file, 'w') as f:
            json.dump(self.inventory, f, indent=2)
    
    def add_item(self, name, quantity, unit, category, expiry_date=None):
        """Add a new item to inventory."""
        # Get next ID
        next_id = max([item['id'] for item in self.inventory], default=0) + 1
        
        item = {
            "id": next_id,
            "name": name,
            "quantity": quantity,
            "unit": unit,
            "category": category,
            "expiry_date": expiry_date if expiry_date else None,
            "added_date": datetime.now().strftime("%Y-%m-%d")
        }
        self.inventory.append(item)
        self.save_inventory()
        return item
    
    def remove_item(self, item_id):
        """Remove an item from inventory by ID."""
        self.inventory = [item for item in self.inventory if item["id"] != item_id]
        self.save_inventory()
        return True
    
    def update_item(self, item_id, **kwargs):
        """Update an item's details."""
        for item in self.inventory:
            if item["id"] == item_id:
                for key, value in kwargs.items():
                    if key in item:
                        item[key] = value
                self.save_inventory()
                return item
        return None
    
    def get_item(self, item_id):
        """Get a specific item by ID."""
        for item in self.inventory:
            if item["id"] == item_id:
                return item
        return None
    
    def get_by_category(self, category):
        """Get all items in a specific category."""
        return [item for item in self.inventory if item["category"].lower() == category.lower()]
    
    def get_all_categories(self):
        """Get all unique categories."""
        categories = set()
        for item in self.inventory:
            categories.add(item["category"])
        return sorted(list(categories))
    
    def export_to_excel(self):
        """Export inventory to Excel file in memory."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        # Define header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        headers = ["ID", "Item Name", "Quantity", "Unit", "Category", "Expiry Date", "Added Date"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Add data rows
        for item in self.inventory:
            ws.append([
                item["id"],
                item["name"],
                item["quantity"],
                item["unit"],
                item["category"],
                item.get("expiry_date", ""),
                item["added_date"]
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# Initialize inventory
inventory = GroceryInventory()


@app.route('/')
def index():
    """Home page - display all items."""
    items = inventory.inventory
    categories = inventory.get_all_categories()
    stats = {
        'total_items': len(items),
        'total_categories': len(categories)
    }
    return render_template('index.html', items=items, categories=categories, stats=stats)


@app.route('/category/<category_name>')
def view_category(category_name):
    """View items by category."""
    items = inventory.get_by_category(category_name)
    categories = inventory.get_all_categories()
    return render_template('category.html', items=items, category=category_name, categories=categories)


@app.route('/add', methods=['GET', 'POST'])
def add_item():
    """Add a new item."""
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            quantity = int(request.form.get('quantity', 1))
            unit = request.form.get('unit', '').strip()
            category = request.form.get('category', '').strip()
            expiry_date = request.form.get('expiry_date', '').strip() or None
            
            if not name or not category or not unit:
                flash('Please fill in all required fields!', 'error')
                return redirect(url_for('add_item'))
            
            inventory.add_item(name, quantity, unit, category, expiry_date)
            flash(f'✓ Item "{name}" added successfully!', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            flash(f'Error adding item: {str(e)}', 'error')
            return redirect(url_for('add_item'))
    
    categories = inventory.get_all_categories()
    return render_template('add_item.html', categories=categories)


@app.route('/edit/<int:item_id>', methods=['GET', 'POST'])
def edit_item(item_id):
    """Edit an item."""
    item = inventory.get_item(item_id)
    
    if not item:
        flash('Item not found!', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            quantity = int(request.form.get('quantity', 1))
            unit = request.form.get('unit', '').strip()
            category = request.form.get('category', '').strip()
            expiry_date = request.form.get('expiry_date', '').strip() or None
            
            if not name or not category or not unit:
                flash('Please fill in all required fields!', 'error')
                return redirect(url_for('edit_item', item_id=item_id))
            
            inventory.update_item(item_id, name=name, quantity=quantity, unit=unit, 
                                 category=category, expiry_date=expiry_date)
            flash(f'✓ Item updated successfully!', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            flash(f'Error updating item: {str(e)}', 'error')
            return redirect(url_for('edit_item', item_id=item_id))
    
    categories = inventory.get_all_categories()
    return render_template('edit_item.html', item=item, categories=categories)


@app.route('/delete/<int:item_id>')
def delete_item(item_id):
    """Delete an item."""
    item = inventory.get_item(item_id)
    
    if not item:
        flash('Item not found!', 'error')
        return redirect(url_for('index'))
    
    item_name = item['name']
    inventory.remove_item(item_id)
    flash(f'✓ Item "{item_name}" deleted successfully!', 'success')
    return redirect(url_for('index'))


@app.route('/export')
def export_excel():
    """Export inventory to Excel."""
    try:
        excel_file = inventory.export_to_excel()
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'grocery_inventory_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        flash(f'Error exporting to Excel: {str(e)}', 'error')
        return redirect(url_for('index'))


@app.route('/clear-all')
def clear_all():
    """Clear all items (with confirmation in template)."""
    inventory.inventory = []
    inventory.save_inventory()
    flash('All items have been cleared!', 'success')
    return redirect(url_for('index'))


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
